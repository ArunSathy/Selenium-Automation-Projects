import openpyxl
from openpyxl import *
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pyperclip
import pandas as pd
from tkinter import filedialog
from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import load_workbook
from tkinter import filedialog



# input_file="C:\\Users\\arhsathy\\Desktop\\Translation_Automation.xlsx"
# excel_data=pd.read_excel(input_file,sheet_name='Translate')   #--- reading the excel sheet using pandas

# loading the excel file using "openpyxl"

base_file=filedialog.askopenfilename(title='Select the Input file',filetypes=(("Excel Files","*.xlsx"),("CSV Files","*.csc"),("All Files","*.*")))
work_book=openpyxl.load_workbook(base_file)
work_sheet=work_book['Translate']

# getting the excel rows & columns

row=work_sheet.max_row
column=work_sheet.max_column

print('Row    : ',row,'\nColumn : ',column)

# appending the data from excel to a new list

input_list=[]
output_list=[]


for i in range(2,row+1):
    list_value=work_sheet.cell(i,1).value
    input_list.append(list_value)
print(input_list)

# assigning the chrome driver path

chromedriver=Service(executable_path='C:\Drivers\chromedriver_win32\chromedriver.exe')

#Opening the Google chrome

browser=webdriver.Chrome(service=chromedriver)
browser.maximize_window()
browser.get('https://translate.google.co.in/')

# doing the operation in website

for i in input_list:
    text_bar=browser.find_element(By.XPATH,'//*[@id="yDmH0d"]/c-wiz/div/div[2]/c-wiz/div[2]/c-wiz/div[1]/div[2]/div[2]/c-wiz[1]/span/span/div/textarea')
    text_bar.send_keys(i)

    time.sleep(2)

    copy_button_xpath='//*[@id="yDmH0d"]/c-wiz/div/div[2]/c-wiz/div[2]/c-wiz/div[1]/div[2]/div[2]/c-wiz[2]/div/div[6]/div/div[6]/div[2]/span[2]/button/div[3]'
    copy_button=WebDriverWait(browser,500).until(EC.presence_of_element_located((By.XPATH,copy_button_xpath)))
    copy_button.click()


    x=pyperclip.paste()
    print('Initial value : ',x)
    output_list.append(x)
    print(output_list)

    text_bar_clear_xpath="/html/body/c-wiz/div/div[2]/c-wiz/div[2]/c-wiz/div[1]/div[2]/div[2]/c-wiz[1]/div[1]/div/div/span/button"
    text_bar_clear=browser.find_element(By.XPATH,text_bar_clear_xpath)
    text_bar_clear.click()

    length = len(output_list)

    for v in range(1, length + 1):
        work_sheet.cell(row=v + 1, column=2, value=output_list[v - 1])
        work_book.save("C:\\Users\\arhsathy\\Desktop\\Translation_Automation.xlsx")

browser.quit()

print('\nSuccessfully Completed...!!!')