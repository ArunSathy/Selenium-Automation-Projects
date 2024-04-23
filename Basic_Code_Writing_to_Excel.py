import time
from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By

driver = webdriver.Chrome(executable_path="C:\\Users\\Public\\Documents\\chromedriver.exe")
wb = openpyxl.load_workbook("C:\\Users\\Public\\Documents\\Sonarname.xlsx")
sh1 = wb['Sheet1']
rows = sh1.max_row
column = sh1.max_column
print(rows, column)
# getpass = os.getlogin()
#print(getpass)
# system_user_name = getpass
# options = webdriver.ChromeOptions()
# options.add_argument('user-data-dir=C:\Users\' + str(system_user_name) + '\AppData\Local\Google\Chrome\User Data')
# driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
# driver.maximize_window()

amazonulr= []
Amazonratings1 = []

for i in range(1, rows + 1):
    x = sh1.cell(i, 1).value
    amazonulr.append(x)
print(amazonulr)

for k in range(1, rows):

    print(amazonulr[k])
    driver.get(amazonulr[k])
    time.sleep(20)

    try:
        review1 = driver.find_element(By.CLASS_NAME, 'displayName').text
        print(review1)
        Amazonratings1.append(review1)
    except Exception:
        review1 = "Not available"
        print(review1)
        Amazonratings1.append(review1)

    length = len(Amazonratings1)

    for v in range(1, length + 1):
        sh1.cell(row=v + 1, column=2, value=Amazonratings1[v - 1])
        wb.save("C:\\Users\\Public\\Documents\\Nameresults.xlsx")