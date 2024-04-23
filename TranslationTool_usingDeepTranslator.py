from deep_translator import GoogleTranslator
import openpyxl
from tkinter import filedialog
import pandas as pd
import os

# user ID

getpass=os.getlogin()
sys_username=getpass

# loading the excel file using "openpyxl"

file_path="C:\\Users\\"+str(sys_username)+"\\Desktop\\Translation_Automation.xlsx"
work_book=openpyxl.load_workbook(file_path)
work_sheet=work_book['Translate']

# getting the excel rows & columns

row=work_sheet.max_row
column=work_sheet.max_column

print('Row    : ',row,'\nColumn : ',column)

# appending the data from excel to a new list

input_list=[]
output_list=[]

for i in range(2,row+1):
    list_value=work_sheet.cell(i,1).value
    input_list.append(list_value)
print(input_list)

# translating each word using a for loop

for i in input_list:
    translator=GoogleTranslator(source='auto',target='english')
    translated_text=translator.translate(i)
    output_list.append(translated_text)

print(output_list)

# saving the output to the same sheet

length = len(output_list)

for j in range(1, length + 1):
    work_sheet.cell(row=j+1, column=2, value=output_list[j-1])
    work_book.save(file_path)

print('\nSuccessfully Executed...!!!')

